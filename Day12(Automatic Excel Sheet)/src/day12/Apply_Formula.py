from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
wb = load_workbook('Barchart.xlsx')
sheet = wb['Report']


min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row


for i in range(min_column+1,max_column+1):
    letter = get_column_letter(i)
    sheet[f'{letter}{max_row+1}']=f'=SUM({letter}{min_row+1}:{letter}{max_row})'
    sheet.column_dimensions[letter].width = 15
    sheet[f'{letter}{max_row+1}'].style = 'Currency'  



wb.save('Report.xlsx')