from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.chart.series import Series
import os 
import sys
application_path = os.path.dirname(sys.executable)
month = input('Introduce Month:')
input_path = os.path.join(application_path, "pivot_table.xlsx")
wb = load_workbook(input_path)
sheet = wb['Report']

min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

chart = BarChart()

chart = BarChart()
chart.type = "col"
chart.title = "Sales by Product Line"
chart.style = 1
chart.y_axis.title = 'Total Sales'
chart.x_axis.title = 'Gender'

data = Reference(sheet, min_col=min_column+1, max_col=max_column, min_row=min_row, max_row=max_row)
categories = Reference(sheet, min_col=min_column, max_col=min_column, min_row=min_row+1, max_row=max_row)
chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)
colors = ['FF5733', '33FF57', '3357FF', 'F1C40F', '9B59B6', '16A085']
for i, ser in enumerate(chart.series):
    ser.graphicalProperties.solidFill = colors[i % len(colors)]



sheet.add_chart(chart, "B12")

for i in range(min_column+1,max_column+1):
    letter = get_column_letter(i)
    sheet[f'{letter}{max_row+1}']=f'=SUM({letter}{min_row+1}:{letter}{max_row})'
    sheet.column_dimensions[letter].width = 15
    sheet[f'{letter}{max_row+1}'].style = 'Currency'

sheet['A1'] = 'Sales Report'
sheet['A2'] = month

sheet['A1'].font = Font('Arial',bold=True, size=20)
sheet['A2'].font = Font('Arial',bold=True, size=10)
output_path = os.path.join(application_path, f'Report_{month}.xlsx')
wb.save(output_path)