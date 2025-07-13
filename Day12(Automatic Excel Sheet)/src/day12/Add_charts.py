from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import series

# Load workbook and 'Report' sheet
wb = load_workbook('pivot_table.xlsx')
sheet = wb['Report']

# Get min and max rows/columns from the 'Report' sheet
min_column = sheet.min_column    # should be 1 (Gender)
max_column = sheet.max_column    # should be 7 (Sports and Travel)
min_row = sheet.min_row          # should be 1 (Header row)
max_row = sheet.max_row          # should be 3 (Female, Male rows)

# Create a BarChart object
chart = BarChart()
chart.type = "col"
chart.title = "Sales by Product Line"
chart.style = 1
chart.y_axis.title = 'Total Sales'
chart.x_axis.title = 'Gender'

# Define the data range for product lines (from 2nd column onwards)
data = Reference(sheet, min_col=min_column+1, max_col=max_column, min_row=min_row, max_row=max_row)

# Define the categories (Gender: Female, Male)
categories = Reference(sheet, min_col=min_column, max_col=min_column, min_row=min_row+1, max_row=max_row)

# Add data and categories to the chart
chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)

# Define custom colors for each product line series
colors = ['FF5733', '33FF57', '3357FF', 'F1C40F', '9B59B6', '16A085']  # HEX RGB colors

# Apply colors to each series
for i, series in enumerate(chart.series):
    series.graphicalProperties.solidFill = colors[i]

# Position the chart on the sheet
sheet.add_chart(chart, "B12")

# Save workbook
wb.save('Barchart.xlsx')
